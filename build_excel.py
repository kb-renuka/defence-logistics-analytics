import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

df = pd.read_csv("cleaned_defence_logistics_data.csv", parse_dates=["Delivery_Date"])

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

wb = Workbook()

# ================= Sheet 1: Cleaned_Data (as a real Excel Table -> PivotTable source) =================
ws = wb.active
ws.title = "Cleaned_Data"
ws["A1"] = "Defence Logistics & Operations — Cleaned Dataset"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Source for PivotTables. Insert > PivotTable in Excel, using this table as the source range."
ws["A2"].font = SUBTITLE_FONT

header_row = 4
cols = list(df.columns)
for c, h in enumerate(cols, start=1):
    ws.cell(row=header_row, column=c, value=h)
style_header_row(ws, header_row, len(cols))

for i, row in df.iterrows():
    rr = header_row + 1 + i
    for c, col in enumerate(cols, start=1):
        val = row[col]
        if col == "Delivery_Date":
            ws.cell(row=rr, column=c, value=val.to_pydatetime() if pd.notna(val) else None).number_format = "dd-mmm-yyyy"
        elif isinstance(val, (bool,)):
            ws.cell(row=rr, column=c, value=str(val))
        else:
            ws.cell(row=rr, column=c, value=val)
        ws.cell(row=rr, column=c).font = BODY_FONT

last_row = header_row + len(df)
last_col_letter = get_column_letter(len(cols))

# Register as a native Excel Table (enables structured refs + is required for a real PivotTable source)
tab = Table(displayName="LogisticsTable", ref=f"A{header_row}:{last_col_letter}{last_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = f"A{header_row+1}"

widths = {get_column_letter(i+1): 16 for i in range(len(cols))}
widths["A"] = 12
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ================= Sheet 2: Lookup_Demo (INDEX/MATCH — Excel-safe alternative to XLOOKUP) =================
lk = wb.create_sheet("Lookup_Demo")
lk["A1"] = "Equipment Reference Lookup (INDEX/MATCH)"
lk["A1"].font = TITLE_FONT
lk["A2"] = "XLOOKUP is avoided here because it doesn't evaluate reliably outside modern Excel/365 in automated pipelines — INDEX/MATCH is the safer, universally-compatible equivalent and is what most enterprises still standardize on."
lk["A2"].font = SUBTITLE_FONT
lk.row_dimensions[2].height = 30
lk.merge_cells("A2:F2")
lk["A2"].alignment = Alignment(wrap_text=True)

ref_headers = ["Equipment_Type", "Typical Unit Cost Tier", "Criticality"]
for c, h in enumerate(ref_headers, start=1):
    lk.cell(row=4, column=c, value=h)
style_header_row(lk, 4, len(ref_headers))

equip_ref = [
    ("Transport Vehicle", "High", "Critical"),
    ("Communication Equipment", "Medium", "Critical"),
    ("Medical Unit", "Medium", "High"),
    ("Artillery Support", "Very High", "Critical"),
    ("Engineering Equipment", "High", "Moderate"),
]
for i, (et, tier, crit) in enumerate(equip_ref):
    rr = 5 + i
    lk.cell(row=rr, column=1, value=et).font = BODY_FONT
    lk.cell(row=rr, column=2, value=tier).font = BODY_FONT
    lk.cell(row=rr, column=3, value=crit).font = BODY_FONT
    for c in range(1, 4):
        lk.cell(row=rr, column=c).border = BORDER

lk["A12"] = "Lookup applied to Cleaned_Data (first 15 rows shown):"
lk["A12"].font = Font(name="Arial", bold=True, size=10)
lk.cell(row=13, column=1, value="Equipment_Type")
lk.cell(row=13, column=2, value="Criticality (INDEX/MATCH)")
style_header_row(lk, 13, 2)
for i in range(15):
    rr = 14 + i
    src_row = header_row + 1 + i
    lk.cell(row=rr, column=1, value=f"=Cleaned_Data!C{src_row}").font = BODY_FONT
    lk.cell(row=rr, column=2,
            value=f'=IFERROR(INDEX($C$5:$C$9,MATCH(A{rr},$A$5:$A$9,0)),"Not in reference table")').font = BODY_FONT
    for c in range(1, 3):
        lk.cell(row=rr, column=c).border = BORDER

lk.column_dimensions["A"].width = 26
lk.column_dimensions["B"].width = 24
lk.column_dimensions["C"].width = 16

# ================= Sheet 3: Analysis (SUMIFS-based summary — the formula layer behind any PivotTable) =================
an = wb.create_sheet("Analysis")
an["A1"] = "Regional & Monthly Summary"
an["A1"].font = TITLE_FONT
an["A2"] = "Formula-driven summary (SUMIFS/COUNTIFS/AVERAGEIFS) that mirrors what a PivotTable on Cleaned_Data would show — recalculates automatically if the data changes."
an["A2"].font = SUBTITLE_FONT
an.row_dimensions[2].height = 28
an.merge_cells("A2:F2")
an["A2"].alignment = Alignment(wrap_text=True)

regions = sorted(df["Region"].unique().tolist())
hdr = 4
an.cell(row=hdr, column=1, value="Region")
an.cell(row=hdr, column=2, value="Total Cost (INR)")
an.cell(row=hdr, column=3, value="Avg Readiness")
an.cell(row=hdr, column=4, value="Delayed Deliveries")
an.cell(row=hdr, column=5, value="Records")
style_header_row(an, hdr, 5)

region_col_letter = None
for idx, colname in enumerate(cols, start=1):
    if colname == "Region":
        region_col_letter = get_column_letter(idx)
    if colname == "Monthly_Cost_INR":
        cost_col_letter = get_column_letter(idx)
    if colname == "Operational_Readiness":
        readiness_col_letter = get_column_letter(idx)
    if colname == "Delivery_Status":
        status_col_letter = get_column_letter(idx)

for i, region in enumerate(regions):
    rr = hdr + 1 + i
    an.cell(row=rr, column=1, value=region).font = BODY_FONT
    an.cell(row=rr, column=2,
            value=f"=SUMIFS(Cleaned_Data!${cost_col_letter}${header_row+1}:${cost_col_letter}${last_row},"
                  f"Cleaned_Data!${region_col_letter}${header_row+1}:${region_col_letter}${last_row},A{rr})")
    an.cell(row=rr, column=2).number_format = '"₹"#,##0'
    an.cell(row=rr, column=3,
            value=f"=AVERAGEIFS(Cleaned_Data!${readiness_col_letter}${header_row+1}:${readiness_col_letter}${last_row},"
                  f"Cleaned_Data!${region_col_letter}${header_row+1}:${region_col_letter}${last_row},A{rr})")
    an.cell(row=rr, column=3).number_format = "0.0%"
    an.cell(row=rr, column=4,
            value=f"=COUNTIFS(Cleaned_Data!${region_col_letter}${header_row+1}:${region_col_letter}${last_row},A{rr},"
                  f"Cleaned_Data!${status_col_letter}${header_row+1}:${status_col_letter}${last_row},\"Delayed\")")
    an.cell(row=rr, column=5,
            value=f"=COUNTIF(Cleaned_Data!${region_col_letter}${header_row+1}:${region_col_letter}${last_row},A{rr})")
    for c in range(1, 6):
        an.cell(row=rr, column=c).font = BODY_FONT
        an.cell(row=rr, column=c).border = BORDER

an_last = hdr + len(regions)
an.column_dimensions["A"].width = 20
for col in ["B", "C", "D", "E"]:
    an.column_dimensions[col].width = 16

# ================= Sheet 4: Dashboard =================
db = wb.create_sheet("Dashboard")
db["B2"] = "Defence Logistics Dashboard"
db["B2"].font = Font(name="Arial", bold=True, size=16, color="1F3864")
db["B3"] = "Excel layer of the 3-tier project (Python/SQL -> Excel -> Power BI)"
db["B3"].font = SUBTITLE_FONT

bar = BarChart()
bar.title = "Total Cost by Region (INR)"
bar.y_axis.title = "INR"
data_ref = Reference(an, min_col=2, min_row=hdr, max_row=an_last)
cats_ref = Reference(an, min_col=1, min_row=hdr + 1, max_row=an_last)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.height = 9
bar.width = 17
db.add_chart(bar, "B6")

pie = PieChart()
pie.title = "Delayed Deliveries by Region"
pdata = Reference(an, min_col=4, min_row=hdr, max_row=an_last)
pcats = Reference(an, min_col=1, min_row=hdr + 1, max_row=an_last)
pie.add_data(pdata, titles_from_data=True)
pie.set_categories(pcats)
pie.height = 9
pie.width = 12
db.add_chart(pie, "K6")

for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
    db.column_dimensions[col].width = 12

wb._sheets = [ws, lk, an, db]
wb.active = 0
wb.save("Defence_Logistics_Analysis.xlsx")
print("saved workbook")
